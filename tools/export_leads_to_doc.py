from typing import Any
import pandas as pd
import os
from datetime import datetime
from core.tool.tool_class import Tool

class ExportLeadsToDocTool(Tool):
    def __init__(self):
        input_schema = {
            "company_name": {
                "type": "string",
                "required": True,
                "description": "The name of the company"
            },
            "phone_number": {
                "type": "array",
                "required": False,
                "description": "The phone number of the company"
            },
            "email": {
                "type": "array",
                "required": False,
                "description": "The email of the company"
            },
            "website": {
                "type": "string",
                "required": True,
                "description": "The website of the company"
            },
            "location": {
                "type": "string",
                "required": True,
                "description": "The location of the company"
            },
            "company_description": {
                "type": "string",
                "required": True,
                "description": "The description of the company"
            },
            "employees": {
                "type": "string",
                "required": False,
                "description": "Employees whose roles and designation that comes in input their email and phone number"
            }
        }
        super().__init__(name="ExportLeadsToDocTool", description="Export leads to an Excel file.", input_schema=input_schema)

    def _get_display_name(self, field_name: str) -> str:
        """
        Convert field name to a nice display name for Excel columns.
        """
        # Convert snake_case to Title Case
        display_name = field_name.replace('_', ' ').title()
        
        # Handle common field name mappings
        field_mappings = {
            'company_name': 'Company Name',
            'phone_number': 'Phone Numbers',
            'email': 'Email Addresses',
            'website': 'Website',
            'location': 'Location',
            'company_description': 'Company Description',
            'employees': 'Employees'
        }
        
        return field_mappings.get(field_name, display_name)
    
    def run(self, input_data: Any) -> Any:
        """
        Export leads data to an Excel file and save it to the local system.
        
        Args:
            input_data: List of lead dictionaries containing company information
            
        Returns:
            Dictionary with success status and file path information
        """
        try:
            # Validate input data first
            validation_result = self.validate_input_schema(input_data)
            if validation_result != True:
                return {
                    "success": False,
                    "server_error": False,
                    "error": f"Input validation failed: {validation_result}"
                }
            
            # Create a directory for exports if it doesn't exist
            export_dir = "exports"
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"leads_export_{timestamp}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            # Prepare data for Excel export dynamically based on schema
            excel_data = []
            for lead in input_data:
                excel_row = {}
                
                for field_name in self.input_schema.keys():
                    if field_name in lead:
                        field_value = lead[field_name]
                        
                        # Handle array fields by joining with line breaks for better readability
                        if isinstance(field_value, list):
                            if field_value:  # Only process non-empty lists
                                display_value = "\n".join(str(item) for item in field_value)
                            else:
                                display_value = ""  # Empty list shows as empty string
                        else:
                            display_value = str(field_value) if field_value is not None else ""
                        
                        # Create a nice display name for the column
                        display_name = self._get_display_name(field_name)
                        excel_row[display_name] = display_value
                
                excel_data.append(excel_row)
            
            # Create DataFrame and export to Excel
            df = pd.DataFrame(excel_data)
            
            # Export to Excel with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Leads', index=False)
                
                # Get the worksheet for formatting
                worksheet = writer.sheets['Leads']
                
                # Auto-adjust column widths based on maximum content length
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    # Check header length first
                    header_cell = worksheet[f"{column_letter}1"]
                    if header_cell.value:
                        max_length = len(str(header_cell.value))
                    
                    # Check all data cells in the column
                    for cell in column[1:]:  # Skip header row
                        try:
                            if cell.value is not None:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except (TypeError, AttributeError):
                            pass
                    
                    # Set width with padding and reasonable limits
                    # Minimum width of 10, maximum of 80, add 3 characters for padding
                    adjusted_width = max(10, min(max_length + 3, 80))
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add comprehensive formatting for better readability
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Header formatting
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Data cell formatting
                data_font = Font(size=10)
                data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Border styling
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Apply header formatting
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Apply data formatting to all data rows
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = thin_border
                
                # Set row height for better readability
                for row in range(1, worksheet.max_row + 1):
                    worksheet.row_dimensions[row].height = 20
            
            # Get absolute file path
            absolute_filepath = os.path.abspath(filepath)
            
            return {
                "success": True,
                "message": f"Successfully exported {len(input_data)} leads to Excel file",
                "filepath": absolute_filepath,
                "filename": filename,
                "total_leads": len(input_data)
            }
            
        except (ValueError, TypeError, KeyError, OSError, PermissionError) as e:
            return {
                "success": False,
                "error": f"Failed to export leads to Excel: {str(e)}"
            }